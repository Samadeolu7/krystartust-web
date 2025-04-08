import csv
from openpyxl import load_workbook


def covert_excel_to_csv(input_file, output_file):
    """
    Convert an Excel file to CSV format.
    :param input_file: Path to the input Excel file.
    :param output_file: Path to the output CSV file.
    """
    # Check if the input file is an Excel file
    if not input_file.endswith('.xlsx'):
        raise ValueError("Input file must be an Excel (.xlsx) file.")

    try:
        # Load the Excel workbook
        workbook = load_workbook(filename=input_file)
        sheet = workbook.active  # Get the active sheet

        # Open the CSV file for writing
        with open(output_file, mode='w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            # Write each row from the Excel sheet to the CSV file
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Successfully converted {input_file} to {output_file}.")
    except Exception as e:
        print(f"An error occurred: {e}")